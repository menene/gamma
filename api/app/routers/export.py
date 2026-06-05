import io
import json
import logging
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.auth import get_current_user
from app.db import get_db

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/export", tags=["Export"], dependencies=[Depends(get_current_user)])


# --- Schemas ---

class RequestPreview(BaseModel):
    id: int
    conversation_id: str | None
    short_text: str | None
    long_text: str | None
    material_type_code: str | None
    category: str | None
    class_name: str | None
    confidence: float | None
    corrected: bool
    status: str
    created_at: datetime
    confirmed_at: datetime | None
    exported_at: datetime | None


class ExportFilters(BaseModel):
    status: list[str] = ["confirmed"]
    exclude_exported: bool = True
    ids: list[int] | None = None


# --- Preview: list requests with filters ---

@router.get("/requests", response_model=list[RequestPreview])
def list_exportable_requests(
    status: str = Query("confirmed", description="Comma-separated statuses"),
    exclude_exported: bool = Query(True),
    db: Session = Depends(get_db),
):
    statuses = [s.strip() for s in status.split(",") if s.strip()]

    where_clauses = []
    params: dict = {}

    if statuses:
        placeholders = ", ".join(f":s{i}" for i in range(len(statuses)))
        where_clauses.append(f"r.status IN ({placeholders})")
        for i, s in enumerate(statuses):
            params[f"s{i}"] = s

    if exclude_exported:
        where_clauses.append("r.exported_at IS NULL")

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    rows = db.execute(
        text(f"""
            SELECT r.id, r.conversation_id, r.short_text, r.long_text,
                   mt.code AS material_type_code,
                   r.category, c.name AS class_name,
                   r.confidence, r.corrected, r.status,
                   r.created_at, r.confirmed_at, r.exported_at
            FROM silver.requests r
            LEFT JOIN silver.material_types mt ON mt.id = r.material_type_id
            LEFT JOIN silver.classes c ON c.code = r.category
            {where_sql}
            ORDER BY r.confirmed_at DESC NULLS LAST, r.created_at DESC
            LIMIT 500
        """),
        params,
    ).fetchall()

    return [
        RequestPreview(
            id=r[0], conversation_id=str(r[1]) if r[1] else None,
            short_text=r[2], long_text=r[3], material_type_code=r[4],
            category=r[5], class_name=r[6],
            confidence=float(r[7]) if r[7] else None,
            corrected=r[8], status=r[9],
            created_at=r[10], confirmed_at=r[11], exported_at=r[12],
        )
        for r in rows
    ]


# --- Export to XLSX ---

class ExportRequest(BaseModel):
    ids: list[int]


class ExportResult(BaseModel):
    exported_count: int
    file_name: str


@router.post("/xlsx")
def export_xlsx(body: ExportRequest, db: Session = Depends(get_db)):
    if not body.ids:
        return StreamingResponse(
            io.BytesIO(b""),
            status_code=400,
            headers={"Content-Type": "application/json"},
        )

    placeholders = ", ".join(f":id{i}" for i in range(len(body.ids)))
    params = {f"id{i}": rid for i, rid in enumerate(body.ids)}

    rows = db.execute(
        text(f"""
            SELECT r.id, r.short_text, r.long_text,
                   mt.code AS material_type_code, mt.description AS material_type_desc,
                   r.category, c.name AS class_name,
                   r.confidence, r.corrected, r.status,
                   r.created_at, r.confirmed_at,
                   r.llm_elapsed_s, r.duplicates_elapsed_s, r.predict_elapsed_s,
                   r.processing_time_s
            FROM silver.requests r
            LEFT JOIN silver.material_types mt ON mt.id = r.material_type_id
            LEFT JOIN silver.classes c ON c.code = r.category
            WHERE r.id IN ({placeholders})
            ORDER BY r.id
        """),
        params,
    ).fetchall()

    # Build XLSX with openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes GAMMA"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "ID", "Descripcion Corta", "Descripcion Larga",
        "Tipo Material", "Tipo Descripcion",
        "Clase", "Clase Descripcion",
        "Confianza", "Corregido", "Estado",
        "Creado", "Confirmado",
        "T. LLM (s)", "T. Duplicados (s)", "T. Prediccion (s)", "T. Total (s)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, r in enumerate(rows, 2):
        values = [
            r[0], r[1], r[2],
            r[3], r[4],
            r[5], r[6],
            float(r[7]) if r[7] else None,
            "Si" if r[8] else "No",
            r[9],
            r[10].strftime("%Y-%m-%d %H:%M") if r[10] else None,
            r[11].strftime("%Y-%m-%d %H:%M") if r[11] else None,
            float(r[12]) if r[12] else None,
            float(r[13]) if r[13] else None,
            float(r[14]) if r[14] else None,
            float(r[15]) if r[15] else None,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=col in (2, 3))

    # Column widths
    widths = [6, 40, 50, 12, 25, 12, 30, 10, 10, 12, 16, 16, 10, 12, 12, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Mark as exported
    db.execute(
        text(f"UPDATE silver.requests SET exported_at = now() WHERE id IN ({placeholders})"),
        params,
    )
    db.commit()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"gamma_export_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
